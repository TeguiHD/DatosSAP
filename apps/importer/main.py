from __future__ import annotations

import argparse
import json
from dataclasses import asdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from parsers.kks_parser import KksDryRun, parse_kks_file
from parsers.plans_parser import PlansDryRun, parse_plans_file
from parsers.positions_parser import (
    PositionsDryRun,
    parse_positions_file,
)


def detect_file(path: Path) -> str:
    workbook = load_workbook(path, read_only=True, data_only=True)
    names = {name.lower() for name in workbook.sheetnames}
    if "actividades mp essc sur" in names:
        return "POSICIONES_ESSC_SUR"
    if "planes" in names:
        return "PLANES_MANTENCION"
    if any("kks essc" in name for name in names):
        return "KKS_FIORI"
    raise ValueError(f"Unsupported Excel structure: {path.name}")


def preview_rows(path: Path, limit: int = 10) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    detected = detect_file(path)
    if detected == "KKS_FIORI":
        sheet_name = next(name for name in workbook.sheetnames if "kks essc" in name.lower())
    elif detected == "POSICIONES_ESSC_SUR":
        sheet_name = "Actividades MP ESSC Sur"
    elif detected == "PLANES_MANTENCION":
        sheet_name = "Planes" if "Planes" in workbook.sheetnames else workbook.sheetnames[0]
    else:
        sheet_name = workbook.sheetnames[0]

    sheet = workbook[sheet_name]
    rows: list[list[Any]] = []
    for row in sheet.iter_rows(values_only=True):
        values = [cell.isoformat() if hasattr(cell, "isoformat") else cell for cell in row[:12]]
        if any(value not in (None, "") for value in values):
            rows.append(values)
        if len(rows) >= limit:
            break

    headers = [str(value) if value not in (None, "") else f"Columna {index + 1}" for index, value in enumerate(rows[0] if rows else [])]
    return {
        "fileType": detected,
        "sheet": sheet_name,
        "headers": headers,
        "rows": rows[1:] if rows else [],
    }


def parse_kks(path: Path) -> KksDryRun:
    return parse_kks_file(path).dry_run()


def parse_posiciones(path: Path) -> PositionsDryRun:
    return parse_positions_file(path).dry_run()


def parse_planes(path: Path) -> PlansDryRun:
    return parse_plans_file(path).dry_run()


def export_kks(path: Path) -> dict[str, Any]:
    return parse_kks_file(path).export_payload()


def export_posiciones(path: Path) -> dict[str, Any]:
    return parse_positions_file(path).export_payload()


def export_planes(path: Path) -> dict[str, Any]:
    return parse_plans_file(path).export_payload()


def export_rows(path: Path, file_type: str | None) -> dict[str, Any]:
    detected = file_type or detect_file(path)
    if detected == "KKS_FIORI":
        return export_kks(path)
    if detected == "POSICIONES_ESSC_SUR":
        return export_posiciones(path)
    if detected == "PLANES_MANTENCION":
        return export_planes(path)
    raise ValueError(f"Unsupported file type: {detected}")


def to_json(
    result: KksDryRun | PositionsDryRun | PlansDryRun,
) -> str:
    payload = asdict(result)
    return json.dumps(payload, ensure_ascii=False, indent=2, default=str)


def dry_run(
    path: Path,
    file_type: str | None,
) -> KksDryRun | PositionsDryRun | PlansDryRun:
    detected = file_type or detect_file(path)
    if detected == "KKS_FIORI":
        return parse_kks(path)
    if detected == "POSICIONES_ESSC_SUR":
        return parse_posiciones(path)
    if detected == "PLANES_MANTENCION":
        return parse_planes(path)
    raise ValueError(f"Unsupported file type: {detected}")


def main() -> None:
    parser = argparse.ArgumentParser(description="datos.nicoholas Excel importer")
    sub = parser.add_subparsers(dest="command", required=True)
    dry = sub.add_parser("dry-run")
    dry.add_argument("--file", required=True)
    dry.add_argument("--type", choices=["KKS_FIORI", "POSICIONES_ESSC_SUR", "PLANES_MANTENCION"])
    export = sub.add_parser("export")
    export.add_argument("--file", required=True)
    export.add_argument("--type", choices=["KKS_FIORI", "POSICIONES_ESSC_SUR", "PLANES_MANTENCION"])
    preview = sub.add_parser("preview")
    preview.add_argument("--file", required=True)
    preview.add_argument("--limit", type=int, default=10)
    args = parser.parse_args()

    path = Path(args.file).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(path)
    if args.command == "dry-run":
        print(to_json(dry_run(path, args.type)))
    if args.command == "export":
        print(json.dumps(export_rows(path, args.type), ensure_ascii=False, indent=2, default=str))
    if args.command == "preview":
        print(json.dumps(preview_rows(path, args.limit), ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()

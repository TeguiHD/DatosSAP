from __future__ import annotations

import argparse
import hashlib
import json
import re
import unicodedata
from collections import Counter
from dataclasses import asdict, dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


FILE_TYPE = "KKS_FIORI"
SHEET_NAME = "KKS ESSC General"

EXPECTED_HEADERS = {
    "node_class": "clase de objeto tecnico",
    "parent_object": "objeto tecnico superior",
    "technical_object": "objeto tecnico",
    "planning_group": "grupo de planificacion",
    "site": "emplazamiento",
    "work_center": "puesto de trabajo principal",
    "system_status": "estado del sistema",
    "location_center": "centro de emplazamiento",
    "cost_center": "centro de coste",
    "wbs_element": "elemento pep",
    "kks": "kks",
    "kks_description": "descripcion kks",
    "equipment": "equipo",
    "equipment_description": "descripcion equipo",
    "center": "centro",
}


@dataclass(frozen=True)
class KksIssue:
    severity: str
    code: str
    message: str
    row_number: int | None = None
    suggested_action: str | None = None


@dataclass(frozen=True)
class KksNode:
    row_number: int
    technical_object: str
    superior_object: str | None
    parent_equipment_code: str | None
    node_type: str
    plant_code: str | None
    kks: str | None
    kks_description: str | None
    equipment_code: str
    equipment_description: str | None
    planning_group: str | None
    site: str | None
    system_status: str | None
    center: str | None
    work_center: str | None
    cost_center: str | None
    wbs_element: str | None
    location_center: str | None
    raw: dict[str, Any]
    source_hash: str
    idempotency_hash: str

    def to_export_dict(self) -> dict[str, Any]:
        return {
            "rowNumber": self.row_number,
            "technicalObject": self.technical_object,
            "superiorObject": self.superior_object,
            "parentEquipmentCode": self.parent_equipment_code,
            "nodeType": self.node_type,
            "plantCode": self.plant_code,
            "kks": self.kks,
            "kksDescription": self.kks_description,
            "equipmentCode": self.equipment_code,
            "equipmentDescription": self.equipment_description,
            "planningGroup": self.planning_group,
            "site": self.site,
            "systemStatus": self.system_status,
            "center": self.center,
            "workCenter": self.work_center,
            "costCenter": self.cost_center,
            "wbsElement": self.wbs_element,
            "locationCenter": self.location_center,
            "raw": self.raw,
            "sourceHash": self.source_hash,
            "idempotencyHash": self.idempotency_hash,
        }


@dataclass(frozen=True)
class KksDryRun:
    file_type: str
    created: int
    updated: int
    skipped: int
    errors: int
    issues: list[KksIssue]
    metadata: dict[str, Any]


@dataclass(frozen=True)
class KksParseResult:
    nodes: list[KksNode]
    issues: list[KksIssue]
    metadata: dict[str, Any]

    def dry_run(self) -> KksDryRun:
        return KksDryRun(
            file_type=FILE_TYPE,
            created=len(self.nodes),
            updated=0,
            skipped=int(self.metadata["skipped_rows"]),
            errors=sum(issue.severity == "CRITICAL" for issue in self.issues),
            issues=self.issues,
            metadata=self.metadata,
        )

    def export_payload(self) -> dict[str, Any]:
        return {
            "fileType": FILE_TYPE,
            "rows": [node.to_export_dict() for node in self.nodes],
        }


def normalize_header(value: Any) -> str:
    text = "" if value is None else str(value).strip()
    decomposed = unicodedata.normalize("NFKD", text)
    without_accents = "".join(character for character in decomposed if not unicodedata.combining(character))
    return " ".join(without_accents.lower().split())


def text_or_none(value: Any) -> str | None:
    if value in (None, ""):
        return None
    text = str(value).strip()
    return text or None


def json_value(value: Any) -> Any:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def stable_hash(payload: Any) -> str:
    serialized = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str)
    return hashlib.sha256(serialized.encode("utf-8")).hexdigest()


def extract_last_parenthetical_code(value: str | None) -> str | None:
    if not value:
        return None
    matches = re.findall(r"\(([^()]*)\)", value)
    return matches[-1].strip() if matches else value.strip()


def plant_code_from_equipment(equipment_code: str, center: str | None) -> str | None:
    upper = equipment_code.upper()
    if upper.startswith(("ESZS-", "ESZN-")):
        return "-".join(upper.split("-")[:2])
    if upper.startswith("EGZN"):
        return "EGZN"
    if upper in {"ESZS", "ESZN"}:
        return upper
    return center.upper() if center else None


def parse_node_type(value: str | None) -> str:
    normalized = normalize_header(value)
    return "EQUIPMENT" if normalized == "equipo" else "TECHNICAL_LOCATION"


def is_non_empty_row(values: Iterable[Any]) -> bool:
    return any(value not in (None, "") for value in values)


def parse_kks_file(path: Path) -> KksParseResult:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        return KksParseResult(
            nodes=[],
            issues=[
                KksIssue(
                    severity="CRITICAL",
                    code="KKS_SHEET_MISSING",
                    message=f"No se encontro la hoja requerida: {SHEET_NAME}.",
                    suggested_action="Verifica que el archivo corresponda a la exportacion KKS Fiori.",
                )
            ],
            metadata={
                "sheet": None,
                "rows": 0,
                "skipped_rows": 0,
                "technical_locations": 0,
                "equipment": 0,
                "root_nodes": 0,
                "resolved_parent_refs": 0,
                "missing_parent_refs": 0,
            },
        )

    sheet = workbook[SHEET_NAME]
    rows = sheet.iter_rows(values_only=True)
    raw_headers = list(next(rows, ()))
    normalized_headers = [normalize_header(value) for value in raw_headers]
    header_index = {header: index for index, header in enumerate(normalized_headers) if header}
    missing_headers = [
        expected
        for expected in EXPECTED_HEADERS.values()
        if expected not in header_index
    ]
    if missing_headers:
        issues = [
            KksIssue(
                severity="CRITICAL",
                code="KKS_HEADER_MISSING",
                message=f"Falta la columna requerida: {header}.",
                suggested_action="Revisa la estructura de la exportacion Fiori antes de aplicar.",
            )
            for header in missing_headers
        ]
        return KksParseResult(
            nodes=[],
            issues=issues,
            metadata={
                "sheet": sheet.title,
                "rows": 0,
                "skipped_rows": 0,
                "technical_locations": 0,
                "equipment": 0,
                "root_nodes": 0,
                "resolved_parent_refs": 0,
                "missing_parent_refs": 0,
            },
        )

    index = {
        field: header_index[expected]
        for field, expected in EXPECTED_HEADERS.items()
    }
    nodes: list[KksNode] = []
    issues: list[KksIssue] = []
    skipped_rows = 0

    for row_number, row in enumerate(rows, start=2):
        if not is_non_empty_row(row):
            continue

        technical_object = text_or_none(row[index["technical_object"]])
        equipment_code = text_or_none(row[index["equipment"]])
        if not technical_object or not equipment_code:
            skipped_rows += 1
            issues.append(
                KksIssue(
                    severity="CRITICAL",
                    code="KKS_IDENTITY_MISSING",
                    message="La fila no contiene Objeto tecnico o Equipo, por lo que no puede importarse de forma idempotente.",
                    row_number=row_number,
                    suggested_action="Corrige la fila en el archivo fuente y vuelve a analizar.",
                )
            )
            continue

        superior_object = text_or_none(row[index["parent_object"]])
        parent_equipment_code = extract_last_parenthetical_code(superior_object)
        if parent_equipment_code == equipment_code:
            parent_equipment_code = None

        raw = {
            (str(raw_headers[position]).strip() if raw_headers[position] is not None else f"col_{position}"): json_value(value)
            for position, value in enumerate(row)
        }
        operational_payload = {
            "technicalObject": technical_object,
            "superiorObject": superior_object,
            "parentEquipmentCode": parent_equipment_code,
            "nodeType": parse_node_type(text_or_none(row[index["node_class"]])),
            "kks": text_or_none(row[index["kks"]]),
            "kksDescription": text_or_none(row[index["kks_description"]]),
            "equipmentCode": equipment_code,
            "equipmentDescription": text_or_none(row[index["equipment_description"]]),
            "planningGroup": text_or_none(row[index["planning_group"]]),
            "site": text_or_none(row[index["site"]]),
            "systemStatus": text_or_none(row[index["system_status"]]),
            "center": text_or_none(row[index["center"]]),
            "workCenter": text_or_none(row[index["work_center"]]),
            "costCenter": text_or_none(row[index["cost_center"]]),
            "wbsElement": text_or_none(row[index["wbs_element"]]),
            "locationCenter": text_or_none(row[index["location_center"]]),
        }
        description = operational_payload["equipmentDescription"] or operational_payload["kksDescription"] or ""
        nodes.append(
            KksNode(
                row_number=row_number,
                technical_object=technical_object,
                superior_object=superior_object,
                parent_equipment_code=parent_equipment_code,
                node_type=str(operational_payload["nodeType"]),
                plant_code=plant_code_from_equipment(
                    equipment_code,
                    text_or_none(row[index["center"]]),
                ),
                kks=text_or_none(row[index["kks"]]),
                kks_description=text_or_none(row[index["kks_description"]]),
                equipment_code=equipment_code,
                equipment_description=text_or_none(row[index["equipment_description"]]),
                planning_group=text_or_none(row[index["planning_group"]]),
                site=text_or_none(row[index["site"]]),
                system_status=text_or_none(row[index["system_status"]]),
                center=text_or_none(row[index["center"]]),
                work_center=text_or_none(row[index["work_center"]]),
                cost_center=text_or_none(row[index["cost_center"]]),
                wbs_element=text_or_none(row[index["wbs_element"]]),
                location_center=text_or_none(row[index["location_center"]]),
                raw=raw,
                source_hash=stable_hash(operational_payload),
                idempotency_hash=stable_hash(
                    {
                        "equipmentCode": equipment_code,
                        "description": description,
                    }
                ),
            )
        )

    technical_objects = Counter(node.technical_object for node in nodes)
    equipment_codes = Counter(node.equipment_code for node in nodes)
    duplicate_technical_objects = [key for key, count in technical_objects.items() if count > 1]
    duplicate_equipment_codes = [key for key, count in equipment_codes.items() if count > 1]
    for field, duplicates in (
        ("Objeto tecnico", duplicate_technical_objects),
        ("Equipo", duplicate_equipment_codes),
    ):
        if duplicates:
            issues.append(
                KksIssue(
                    severity="CRITICAL",
                    code="KKS_IDENTITY_DUPLICATED",
                    message=f"{field} contiene {len(duplicates)} claves duplicadas.",
                    suggested_action="Resuelve los duplicados antes de aplicar la importacion.",
                )
            )

    available_equipment_codes = set(equipment_codes)
    missing_parent_nodes = [
        node
        for node in nodes
        if node.parent_equipment_code
        and node.parent_equipment_code not in available_equipment_codes
    ]
    for node in missing_parent_nodes[:20]:
        issues.append(
            KksIssue(
                severity="WARNING",
                code="KKS_PARENT_NOT_FOUND",
                message=f"No se encontro el padre SAP {node.parent_equipment_code}.",
                row_number=node.row_number,
                suggested_action="El nodo puede importarse como raiz temporal, pero requiere revision de jerarquia.",
            )
        )

    node_types = Counter(node.node_type for node in nodes)
    centers = Counter(node.center or "SIN_CENTRO" for node in nodes)
    kks_codes = Counter(node.kks for node in nodes if node.kks)
    root_nodes = [node for node in nodes if not node.parent_equipment_code]
    resolved_parent_refs = sum(
        bool(node.parent_equipment_code and node.parent_equipment_code in available_equipment_codes)
        for node in nodes
    )

    return KksParseResult(
        nodes=nodes,
        issues=issues,
        metadata={
            "sheet": sheet.title,
            "rows": len(nodes),
            "skipped_rows": skipped_rows,
            "technical_locations": node_types["TECHNICAL_LOCATION"],
            "equipment": node_types["EQUIPMENT"],
            "root_nodes": len(root_nodes),
            "resolved_parent_refs": resolved_parent_refs,
            "missing_parent_refs": len(missing_parent_nodes),
            "unique_technical_objects": len(technical_objects),
            "unique_equipment_codes": len(equipment_codes),
            "unique_kks_codes": len(kks_codes),
            "duplicate_kks_rows": sum(count - 1 for count in kks_codes.values() if count > 1),
            "centers": dict(centers.most_common()),
            "identity_field": "technicalObject",
            "parent_lookup_field": "equipmentCode",
        },
    )


def to_json(payload: Any) -> str:
    if hasattr(payload, "__dataclass_fields__"):
        payload = asdict(payload)
    return json.dumps(payload, ensure_ascii=False, indent=2, default=str)


def main() -> None:
    parser = argparse.ArgumentParser(description="Parser industrial del arbol KKS Fiori")
    parser.add_argument("--file", required=True, help="Ruta al archivo Excel KKS Fiori")
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument("--dry-run", action="store_true", help="Valida y resume sin escribir datos")
    mode.add_argument("--export", action="store_true", help="Exporta las filas normalizadas como JSON")
    args = parser.parse_args()

    path = Path(args.file).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(path)

    result = parse_kks_file(path)
    if args.export:
        print(to_json(result.export_payload()))
        return
    print(to_json(result.dry_run()))


if __name__ == "__main__":
    main()

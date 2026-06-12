from __future__ import annotations

import argparse
import calendar
import json
import re
from collections import Counter
from dataclasses import asdict, dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

try:
    from parsers.kks_parser import (
        extract_last_parenthetical_code,
        json_value,
        normalize_header,
        stable_hash,
        text_or_none,
    )
except ModuleNotFoundError:
    from kks_parser import (
        extract_last_parenthetical_code,
        json_value,
        normalize_header,
        stable_hash,
        text_or_none,
    )


FILE_TYPE = "POSICIONES_ESSC_SUR"
SHEET_NAME = "Actividades MP ESSC Sur"

FREQUENCY_MAP = {
    "1M": ("ONE_MONTH", 1),
    "6M": ("SIX_MONTHS", 6),
    "1A": ("ONE_YEAR", 12),
    "5A": ("FIVE_YEARS", 60),
}

# The final legacy block omits technical location and equipment. These exact
# plant names are present in the activity text and map unambiguously to the
# canonical KKS roots validated in the preceding KKS checkpoint.
LEGACY_ACTIVITY_PLANT_HINTS = {
    "ALIFRUT": "ESZS-60",
    "CCU": "ESZS-50",
    "GOLDEN": "ESZS-10",
    "EDEN": "ESZS-A0",
    "GOODYEAR": "ESZS-90",
    "SIKA": "ESZS-A3",
    "MYLPAN": "ESZS-80",
    "CEMIN": "ESZS-A1",
}


@dataclass(frozen=True)
class PositionsIssue:
    severity: str
    code: str
    message: str
    row_number: int | None = None
    suggested_action: str | None = None


@dataclass(frozen=True)
class MaintenanceOccurrenceRow:
    scheduled_date: date
    due_date: date
    is_historical: bool
    source_month_key: str
    source_value: str
    source_hash: str

    def to_export_dict(self) -> dict[str, Any]:
        return {
            "scheduledFor": self.scheduled_date.isoformat(),
            "dueDate": self.due_date.isoformat(),
            "isHistorical": self.is_historical,
            "sourceMonthKey": self.source_month_key,
            "sourceValue": self.source_value,
            "sourceHash": self.source_hash,
        }


@dataclass(frozen=True)
class MaintenanceTemplateRow:
    row_number: int
    plant_code: str | None
    wbs_element: str | None
    plan_name: str | None
    route_sheet: str | None
    technical_location: str | None
    technical_object: str | None
    equipment: str | None
    equipment_code: str | None
    source_position: str | None
    activity_name: str
    frequency_label: str | None
    frequency: str
    source_frequency: str
    months_interval: int
    start_month: int | None
    context_inferred: bool
    occurrences: list[MaintenanceOccurrenceRow]
    raw: dict[str, Any]
    source_hash: str
    idempotency_hash: str

    def to_export_dict(self) -> dict[str, Any]:
        return {
            "rowNumber": self.row_number,
            "plantCode": self.plant_code,
            "wbsElement": self.wbs_element,
            "planName": self.plan_name,
            "routeSheet": self.route_sheet,
            "technicalLocation": self.technical_location,
            "technicalObject": self.technical_object,
            "equipment": self.equipment,
            "equipmentCode": self.equipment_code,
            "sourcePosition": self.source_position,
            "activityName": self.activity_name,
            "frequencyLabel": self.frequency_label,
            "frequency": self.frequency,
            "sourceFrequency": self.source_frequency,
            "monthsInterval": self.months_interval,
            "startMonth": self.start_month,
            "estimatedHours": None,
            "contextInferred": self.context_inferred,
            "occurrences": [
                occurrence.to_export_dict()
                for occurrence in self.occurrences
            ],
            "raw": self.raw,
            "sourceHash": self.source_hash,
            "idempotencyHash": self.idempotency_hash,
        }


@dataclass(frozen=True)
class PositionsDryRun:
    file_type: str
    created: int
    updated: int
    skipped: int
    errors: int
    issues: list[PositionsIssue]
    metadata: dict[str, Any]


@dataclass(frozen=True)
class PositionsParseResult:
    templates: list[MaintenanceTemplateRow]
    issues: list[PositionsIssue]
    metadata: dict[str, Any]

    def dry_run(self) -> PositionsDryRun:
        return PositionsDryRun(
            file_type=FILE_TYPE,
            created=len(self.templates),
            updated=0,
            skipped=int(self.metadata["skipped_rows"]),
            errors=sum(issue.severity == "CRITICAL" for issue in self.issues),
            issues=self.issues,
            metadata=self.metadata,
        )

    def export_payload(self) -> dict[str, Any]:
        return {
            "fileType": FILE_TYPE,
            "templates": [
                template.to_export_dict()
                for template in self.templates
            ],
        }


class PositionsParser:
    def __init__(
        self,
        conn: Any | None = None,
        import_job_id: str | None = None,
        org_id: str | None = None,
        dry_run: bool = True,
        today: date | None = None,
    ) -> None:
        # Database writes remain in the NestJS transaction. These fields keep
        # the worker contract ready without allowing dry-run to mutate state.
        self.conn = conn
        self.import_job_id = import_job_id
        self.org_id = org_id
        self.dry_run_mode = dry_run
        self.today = today or date.today()

    def run(self, excel_path: Path) -> PositionsParseResult:
        values_workbook = load_workbook(
            excel_path,
            read_only=True,
            data_only=True,
        )
        formula_workbook = load_workbook(
            excel_path,
            read_only=True,
            data_only=False,
        )
        if SHEET_NAME not in values_workbook.sheetnames:
            return self._fatal_result(
                "POSITIONS_SHEET_MISSING",
                f"No se encontro la hoja requerida: {SHEET_NAME}.",
            )

        sheet = values_workbook[SHEET_NAME]
        formula_sheet = formula_workbook[SHEET_NAME]
        header_row = self._find_header_row(sheet)
        if header_row is None:
            return self._fatal_result(
                "POSITIONS_HEADER_NOT_FOUND",
                "No se encontro la fila que contiene Elemento PEP en la columna B.",
            )

        month_columns = self._extract_month_dates(sheet, header_row)
        issues: list[PositionsIssue] = []
        if len(month_columns) != 96:
            issues.append(
                PositionsIssue(
                    severity="CRITICAL",
                    code="POSITIONS_MONTH_COUNT_INVALID",
                    message=f"Se esperaban 96 columnas mensuales y se encontraron {len(month_columns)}.",
                    suggested_action="Verifica que la matriz cubra enero de 2022 a diciembre de 2029.",
                )
            )

        raw_headers = list(
            next(
                sheet.iter_rows(
                    min_row=header_row,
                    max_row=header_row,
                    values_only=True,
                )
            )
        )
        templates: list[MaintenanceTemplateRow] = []
        skipped_rows = 0
        unknown_frequencies: Counter[str] = Counter()
        frequency_mismatches: list[tuple[int, str, int, int]] = []
        inferred_rows: list[int] = []
        missing_asset_context_rows: list[int] = []
        unknown_plant_rows: list[int] = []
        cemin_rows: list[int] = []
        formulas_seen = False

        value_rows = sheet.iter_rows(
            min_row=header_row + 1,
            values_only=True,
        )
        formula_rows = formula_sheet.iter_rows(
            min_row=header_row + 1,
            values_only=True,
        )
        for row_number, (row, formula_row) in enumerate(
            zip(value_rows, formula_rows),
            start=header_row + 1,
        ):
            if not any(value not in (None, "") for value in row):
                continue

            formula_frequency = text_or_none(formula_row[9]) if len(formula_row) > 9 else None
            if formula_frequency and formula_frequency.startswith("="):
                formulas_seen = True

            source_frequency = (
                text_or_none(row[9]).upper()
                if len(row) > 9 and text_or_none(row[9])
                else ""
            )
            frequency_mapping = FREQUENCY_MAP.get(source_frequency)
            if frequency_mapping is None:
                unknown_frequencies[source_frequency or "VACIA"] += 1
                skipped_rows += 1
                continue

            frequency, expected_months = frequency_mapping
            months_interval = self._integer_or_none(
                row[10] if len(row) > 10 else None
            )
            if months_interval is None:
                months_interval = expected_months
            elif months_interval != expected_months:
                frequency_mismatches.append(
                    (
                        row_number,
                        source_frequency,
                        expected_months,
                        months_interval,
                    )
                )

            technical_location = text_or_none(
                row[4] if len(row) > 4 else None
            )
            equipment = text_or_none(
                row[5] if len(row) > 5 else None
            )
            source_position = text_or_none(
                row[6] if len(row) > 6 else None
            )
            activity_name = (
                text_or_none(row[7] if len(row) > 7 else None)
                or source_position
                or f"Mantencion fila {row_number}"
            )
            technical_object = extract_last_parenthetical_code(
                technical_location
            )
            equipment_code = extract_last_parenthetical_code(equipment)
            plant_code = self._resolve_plant_code(
                technical_object,
                activity_name,
            )
            context_inferred = not technical_location and bool(plant_code)
            if context_inferred:
                inferred_rows.append(row_number)
            if not technical_object and not equipment_code:
                missing_asset_context_rows.append(row_number)
            if not plant_code:
                unknown_plant_rows.append(row_number)
            if plant_code == "ESZS-A1":
                cemin_rows.append(row_number)

            idempotency_payload = {
                "plantCode": plant_code,
                "technicalObject": technical_object or "",
                "activityName": activity_name,
                "frequency": frequency,
            }
            idempotency_hash = stable_hash(idempotency_payload)
            occurrences = self._build_occurrences(
                row,
                month_columns,
                idempotency_hash,
            )
            raw = {
                (
                    str(raw_headers[position]).strip()
                    if raw_headers[position] is not None
                    else f"col_{position}"
                ): json_value(value)
                for position, value in enumerate(row)
            }
            operational_payload = {
                **idempotency_payload,
                "equipmentCode": equipment_code,
                "wbsElement": text_or_none(row[1] if len(row) > 1 else None),
                "planName": text_or_none(row[2] if len(row) > 2 else None),
                "routeSheet": text_or_none(row[3] if len(row) > 3 else None),
                "frequencyLabel": text_or_none(row[8] if len(row) > 8 else None),
                "monthsInterval": months_interval,
                "startMonth": self._integer_or_none(
                    row[11] if len(row) > 11 else None
                ),
            }
            templates.append(
                MaintenanceTemplateRow(
                    row_number=row_number,
                    plant_code=plant_code,
                    wbs_element=operational_payload["wbsElement"],
                    plan_name=operational_payload["planName"],
                    route_sheet=operational_payload["routeSheet"],
                    technical_location=technical_location,
                    technical_object=technical_object,
                    equipment=equipment,
                    equipment_code=equipment_code,
                    source_position=source_position,
                    activity_name=activity_name,
                    frequency_label=operational_payload["frequencyLabel"],
                    frequency=frequency,
                    source_frequency=source_frequency,
                    months_interval=months_interval,
                    start_month=operational_payload["startMonth"],
                    context_inferred=context_inferred,
                    occurrences=occurrences,
                    raw=raw,
                    source_hash=stable_hash(operational_payload),
                    idempotency_hash=idempotency_hash,
                )
            )

        if formulas_seen:
            issues.append(
                PositionsIssue(
                    severity="INFO",
                    code="POSITIONS_CACHED_FORMULAS_USED",
                    message="La columna Frec. contiene formulas y se leyeron sus valores cacheados.",
                    suggested_action="Mantener data_only=True para este tipo de archivo.",
                )
            )
        if inferred_rows:
            issues.append(
                PositionsIssue(
                    severity="WARNING",
                    code="POSITIONS_PLANT_CONTEXT_INFERRED",
                    message=f"{len(inferred_rows)} filas legacy no traen ubicacion tecnica; la planta se infirio desde la actividad.",
                    row_number=inferred_rows[0],
                    suggested_action="Completar Ubicacion tecnica y Equipo en la fuente para futuras importaciones.",
                )
            )
        if missing_asset_context_rows:
            issues.append(
                PositionsIssue(
                    severity="WARNING",
                    code="POSITIONS_ASSET_CONTEXT_MISSING",
                    message=f"{len(missing_asset_context_rows)} plantillas no incluyen Equipo ni Ubicacion tecnica y quedaran sin activo asociado.",
                    row_number=missing_asset_context_rows[0],
                    suggested_action="La plantilla puede ser de nivel planta, pero conviene completar el activo en la fuente.",
                )
            )
        if frequency_mismatches:
            issues.append(
                PositionsIssue(
                    severity="WARNING",
                    code="POSITIONS_FREQUENCY_MONTHS_MISMATCH",
                    message=f"{len(frequency_mismatches)} filas tienen una diferencia entre Frec. y Meses; se usara Meses como fuente de verdad.",
                    row_number=frequency_mismatches[0][0],
                    suggested_action="Revisar la consistencia de frecuencia en el Excel.",
                )
            )
        if unknown_frequencies:
            issues.append(
                PositionsIssue(
                    severity="WARNING",
                    code="POSITIONS_FREQUENCY_UNSUPPORTED",
                    message=f"Se omitieron frecuencias no soportadas: {dict(unknown_frequencies)}.",
                    suggested_action="Definir el mapeo antes de volver a analizar.",
                )
            )
        if unknown_plant_rows:
            issues.append(
                PositionsIssue(
                    severity="CRITICAL",
                    code="POSITIONS_PLANT_UNKNOWN",
                    message=f"{len(unknown_plant_rows)} filas no pudieron asociarse a una planta.",
                    row_number=unknown_plant_rows[0],
                    suggested_action="Completa la ubicacion tecnica o define una homologacion de planta.",
                )
            )
        if cemin_rows:
            issues.append(
                PositionsIssue(
                    severity="CRITICAL",
                    code="CEMIN_ALIAS_REQUIRED",
                    message="Se encontro un codigo de planta que no coincide con el catalogo. Selecciona la planta correcta para continuar.",
                    row_number=cemin_rows[0],
                    suggested_action="Resolver la homologacion de planta y volver a analizar.",
                )
            )

        template_hashes = Counter(
            template.idempotency_hash
            for template in templates
        )
        duplicate_template_hashes = [
            key
            for key, count in template_hashes.items()
            if count > 1
        ]
        if duplicate_template_hashes:
            issues.append(
                PositionsIssue(
                    severity="CRITICAL",
                    code="POSITIONS_TEMPLATE_IDENTITY_DUPLICATED",
                    message=f"Se detectaron {len(duplicate_template_hashes)} identidades de plantilla duplicadas.",
                    suggested_action="Revisar planta, actividad y frecuencia antes de aplicar.",
                )
            )

        occurrence_count = sum(
            len(template.occurrences)
            for template in templates
        )
        historical_count = sum(
            occurrence.is_historical
            for template in templates
            for occurrence in template.occurrences
        )
        frequency_counts = Counter(
            template.source_frequency
            for template in templates
        )
        plant_counts = Counter(
            template.plant_code or "SIN_PLANTA"
            for template in templates
        )

        return PositionsParseResult(
            templates=templates,
            issues=issues,
            metadata={
                "sheet": sheet.title,
                "header_row": header_row,
                "month_header_row": header_row,
                "templates": len(templates),
                "month_columns": len(month_columns),
                "occurrences": occurrence_count,
                "historical_occurrences": historical_count,
                "future_occurrences": occurrence_count - historical_count,
                "first_month": (
                    month_columns[0][1].isoformat()
                    if month_columns
                    else None
                ),
                "last_month": (
                    month_columns[-1][1].isoformat()
                    if month_columns
                    else None
                ),
                "frequencies": dict(frequency_counts),
                "plants": dict(plant_counts),
                "context_inferred_rows": len(inferred_rows),
                "asset_context_missing_rows": len(missing_asset_context_rows),
                "cemin_rows": len(cemin_rows),
                "frequency_mismatch_rows": len(frequency_mismatches),
                "unknown_frequency_rows": sum(unknown_frequencies.values()),
                "unknown_plant_rows": len(unknown_plant_rows),
                "skipped_rows": skipped_rows,
            },
        )

    def _find_header_row(self, worksheet: Any) -> int | None:
        for row_number, row in enumerate(
            worksheet.iter_rows(
                min_row=1,
                max_row=min(50, worksheet.max_row),
                values_only=True,
            ),
            start=1,
        ):
            if len(row) <= 1:
                continue
            if "elemento pep" in normalize_header(row[1]):
                return row_number
        return None

    def _extract_month_dates(
        self,
        worksheet: Any,
        header_row: int,
    ) -> list[tuple[int, date]]:
        header = list(
            next(
                worksheet.iter_rows(
                    min_row=header_row,
                    max_row=header_row,
                    values_only=True,
                )
            )
        )
        return [
            (index, value.date())
            for index, value in enumerate(header)
            if index >= 12 and isinstance(value, datetime)
        ]

    def _build_occurrences(
        self,
        row: tuple[Any, ...],
        month_columns: list[tuple[int, date]],
        template_hash: str,
    ) -> list[MaintenanceOccurrenceRow]:
        occurrences: list[MaintenanceOccurrenceRow] = []
        for index, month_date in month_columns:
            value = row[index] if index < len(row) else None
            if value in (None, ""):
                continue
            scheduled_date = date(
                month_date.year,
                month_date.month,
                1,
            )
            due_date = date(
                month_date.year,
                month_date.month,
                calendar.monthrange(
                    month_date.year,
                    month_date.month,
                )[1],
            )
            source_value = str(value).strip()
            occurrences.append(
                MaintenanceOccurrenceRow(
                    scheduled_date=scheduled_date,
                    due_date=due_date,
                    is_historical=scheduled_date < self.today,
                    source_month_key=scheduled_date.strftime("%Y-%m"),
                    source_value=source_value,
                    source_hash=stable_hash(
                        {
                            "templateHash": template_hash,
                            "scheduledDate": scheduled_date.isoformat(),
                        }
                    ),
                )
            )
        return occurrences

    def _resolve_plant_code(
        self,
        technical_object: str | None,
        activity_name: str,
    ) -> str | None:
        if technical_object:
            match = re.match(
                r"^(ESZS-[A-Z0-9]+)",
                technical_object.upper(),
            )
            if match:
                return match.group(1)

        upper_activity = activity_name.upper()
        for hint, plant_code in LEGACY_ACTIVITY_PLANT_HINTS.items():
            if hint in upper_activity:
                return plant_code
        return None

    def _integer_or_none(self, value: Any) -> int | None:
        if isinstance(value, bool):
            return None
        if isinstance(value, (int, float)):
            return int(value)
        if value in (None, ""):
            return None
        try:
            return int(str(value).strip())
        except ValueError:
            return None

    def _fatal_result(
        self,
        code: str,
        message: str,
    ) -> PositionsParseResult:
        return PositionsParseResult(
            templates=[],
            issues=[
                PositionsIssue(
                    severity="CRITICAL",
                    code=code,
                    message=message,
                )
            ],
            metadata={
                "sheet": None,
                "header_row": None,
                "month_header_row": None,
                "templates": 0,
                "month_columns": 0,
                "occurrences": 0,
                "historical_occurrences": 0,
                "future_occurrences": 0,
                "context_inferred_rows": 0,
                "asset_context_missing_rows": 0,
                "cemin_rows": 0,
                "frequency_mismatch_rows": 0,
                "unknown_frequency_rows": 0,
                "unknown_plant_rows": 0,
                "skipped_rows": 0,
            },
        )


def parse_positions_file(
    path: Path,
    today: date | None = None,
) -> PositionsParseResult:
    return PositionsParser(
        dry_run=True,
        today=today,
    ).run(path)


def to_json(payload: Any) -> str:
    if hasattr(payload, "__dataclass_fields__"):
        payload = asdict(payload)
    return json.dumps(
        payload,
        ensure_ascii=False,
        indent=2,
        default=str,
    )


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Parser industrial de Posiciones ESSC Sur"
    )
    parser.add_argument(
        "--file",
        required=True,
        help="Ruta al archivo Excel de Posiciones",
    )
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument(
        "--dry-run",
        action="store_true",
        help="Valida y resume sin escribir datos",
    )
    mode.add_argument(
        "--export",
        action="store_true",
        help="Exporta plantillas y ocurrencias normalizadas",
    )
    args = parser.parse_args()

    path = Path(args.file).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(path)

    result = parse_positions_file(path)
    if args.export:
        print(to_json(result.export_payload()))
        return
    print(to_json(result.dry_run()))


if __name__ == "__main__":
    main()

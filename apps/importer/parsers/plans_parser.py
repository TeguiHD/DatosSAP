from __future__ import annotations

import argparse
import json
import unicodedata
from collections import Counter
from dataclasses import asdict, dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

try:
    from parsers.kks_parser import json_value, stable_hash, text_or_none
except ModuleNotFoundError:
    from kks_parser import json_value, stable_hash, text_or_none


FILE_TYPE = "PLANES_MANTENCION"

REQUIRED_HEADERS = {
    "plan_id",
    "equipo",
    "kks",
    "descripcion_plan",
    "fecha_inicio",
    "fecha_termino",
    "hh_planificadas",
    "hh_ejecutadas",
    "porcentaje_avance",
    "estado",
    "criticidad",
    "puesto_trabajo",
    "especialidad",
    "personal_asignado",
}


@dataclass(frozen=True)
class PlansIssue:
    severity: str
    code: str
    message: str
    row_number: int | None = None
    suggested_action: str | None = None


@dataclass(frozen=True)
class WorkOrderRow:
    row_number: int
    plan_id: str
    work_order_number: str
    equipment_number: str | None
    kks: str | None
    title: str
    scheduled_start_date: date | None
    scheduled_end_date: date | None
    planned_hours: float | None
    actual_hours: float | None
    imported_progress: float | None
    status: str
    source_status: str | None
    criticality: str
    source_criticality: str | None
    required_specialty: str | None
    specialty: str | None
    assigned_to: str | None
    raw: dict[str, Any]
    source_hash: str

    def to_export_dict(self) -> dict[str, Any]:
        metadata = {
            "sourcePlanId": self.plan_id,
            "kks": self.kks,
            "especialidad": self.specialty,
            "personalAsignado": self.assigned_to,
            "sourceStatus": self.source_status,
            "sourceCriticality": self.source_criticality,
        }
        return {
            "rowNumber": self.row_number,
            "planId": self.plan_id,
            "workOrderNumber": self.work_order_number,
            "equipmentNumber": self.equipment_number,
            "equipment": self.equipment_number,
            "kks": self.kks,
            "title": self.title,
            "scheduledStartDate": self.scheduled_start_date.isoformat() if self.scheduled_start_date else None,
            "scheduledEndDate": self.scheduled_end_date.isoformat() if self.scheduled_end_date else None,
            "plannedStart": self.scheduled_start_date.isoformat() if self.scheduled_start_date else None,
            "plannedEnd": self.scheduled_end_date.isoformat() if self.scheduled_end_date else None,
            "plannedHours": self.planned_hours,
            "actualHours": self.actual_hours,
            "importedProgress": self.imported_progress,
            "status": self.status,
            "criticality": self.criticality,
            "requiredSpecialty": self.required_specialty,
            "assignedTo": self.assigned_to,
            "metadata": {key: value for key, value in metadata.items() if value not in (None, "")},
            "raw": self.raw,
            "sourceHash": self.source_hash,
        }


@dataclass(frozen=True)
class PlansDryRun:
    file_type: str
    created: int
    updated: int
    skipped: int
    errors: int
    issues: list[PlansIssue]
    metadata: dict[str, Any]


@dataclass(frozen=True)
class PlansParseResult:
    work_orders: list[WorkOrderRow]
    issues: list[PlansIssue]
    metadata: dict[str, Any]

    def dry_run(self) -> PlansDryRun:
        return PlansDryRun(
            file_type=FILE_TYPE,
            created=len(self.work_orders),
            updated=0,
            skipped=int(self.metadata["skipped_rows"]),
            errors=sum(issue.severity == "CRITICAL" for issue in self.issues),
            issues=self.issues,
            metadata=self.metadata,
        )

    def export_payload(self) -> dict[str, Any]:
        return {
            "fileType": FILE_TYPE,
            "workOrders": [work_order.to_export_dict() for work_order in self.work_orders],
        }


def normalize_header(value: Any) -> str:
    text = "" if value is None else str(value).strip()
    decomposed = unicodedata.normalize("NFKD", text)
    without_accents = "".join(character for character in decomposed if not unicodedata.combining(character))
    return "_".join(without_accents.lower().split())


def is_non_empty_row(values: Iterable[Any]) -> bool:
    return any(value not in (None, "") for value in values)


def raw_row(headers: list[Any], row: Iterable[Any]) -> dict[str, Any]:
    return {
        str(headers[index]).strip() if index < len(headers) and headers[index] not in (None, "") else f"col_{index}": json_value(value)
        for index, value in enumerate(row)
    }


def parse_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    return date.fromisoformat(text[:10])


def parse_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", ".")
    if not text:
        return None
    return float(text)


def map_status(value: str | None) -> str:
    normalized = (value or "").strip().lower()
    if normalized == "completado":
        return "COMPLETED"
    if normalized == "en curso":
        return "IN_PROGRESS"
    return "SCHEDULED"


def map_criticality(value: str | None) -> str:
    normalized = (value or "").strip().lower()
    if normalized in {"critica", "crítica"}:
        return "CRITICAL"
    if normalized == "media":
        return "WARNING"
    if normalized == "baja":
        return "INFO"
    return "CRITICAL"


def find_header_sheet(workbook: Any) -> tuple[Any | None, int | None, list[Any] | None]:
    for sheet in workbook.worksheets:
        for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            headers = [normalize_header(value) for value in row]
            if REQUIRED_HEADERS.issubset(set(headers)):
                return sheet, row_number, list(row)
            if row_number >= 10:
                break
    return None, None, None


def parse_plans_file(path: Path) -> PlansParseResult:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_names = workbook.sheetnames
    sheet, header_row, raw_headers = find_header_sheet(workbook)
    if not sheet or header_row is None or raw_headers is None:
        return PlansParseResult(
            work_orders=[],
            issues=[
                PlansIssue(
                    severity="CRITICAL",
                    code="PLANES_HEADER_MISSING",
                    message="No se encontro una hoja con las columnas requeridas de Planes Mantencion.",
                    suggested_action=f"Hojas disponibles: {', '.join(sheet_names)}",
                )
            ],
            metadata={
                "sheet": None,
                "sheet_names": sheet_names,
                "header_row": None,
                "work_orders": 0,
                "planned_hours": 0,
                "actual_hours": 0,
                "skipped_rows": 0,
            },
        )

    headers = [normalize_header(value) for value in raw_headers]
    header_index = {header: index for index, header in enumerate(headers) if header}
    rows = sheet.iter_rows(min_row=header_row + 1, values_only=True)
    work_orders: list[WorkOrderRow] = []
    issues: list[PlansIssue] = []
    skipped_rows = 0

    for row_number, row in enumerate(rows, start=header_row + 1):
        if not is_non_empty_row(row):
            skipped_rows += 1
            continue

        raw = raw_row(raw_headers, row)
        plan_id = text_or_none(row[header_index["plan_id"]])
        title = text_or_none(row[header_index["descripcion_plan"]])
        if not plan_id or not title:
            issues.append(
                PlansIssue(
                    severity="CRITICAL",
                    code="PLANES_IDENTITY_MISSING",
                    message="La fila no contiene plan_id o descripcion_plan.",
                    row_number=row_number,
                    suggested_action="Corrige la fila en el Excel antes de aplicar.",
                )
            )
            continue

        source_status = text_or_none(row[header_index["estado"]])
        source_criticality = text_or_none(row[header_index["criticidad"]])
        source_payload = {
            "planId": plan_id,
            "equipmentNumber": text_or_none(row[header_index["equipo"]]),
            "title": title,
            "scheduledStartDate": json_value(row[header_index["fecha_inicio"]]),
            "scheduledEndDate": json_value(row[header_index["fecha_termino"]]),
            "plannedHours": json_value(row[header_index["hh_planificadas"]]),
            "actualHours": json_value(row[header_index["hh_ejecutadas"]]),
            "importedProgress": json_value(row[header_index["porcentaje_avance"]]),
            "status": source_status,
            "criticality": source_criticality,
        }
        work_orders.append(
            WorkOrderRow(
                row_number=row_number,
                plan_id=plan_id,
                work_order_number=f"PLAN-{plan_id}",
                equipment_number=text_or_none(row[header_index["equipo"]]),
                kks=text_or_none(row[header_index["kks"]]),
                title=title,
                scheduled_start_date=parse_date(row[header_index["fecha_inicio"]]),
                scheduled_end_date=parse_date(row[header_index["fecha_termino"]]),
                planned_hours=parse_float(row[header_index["hh_planificadas"]]),
                actual_hours=parse_float(row[header_index["hh_ejecutadas"]]),
                imported_progress=parse_float(row[header_index["porcentaje_avance"]]),
                status=map_status(source_status),
                source_status=source_status,
                criticality=map_criticality(source_criticality),
                source_criticality=source_criticality,
                required_specialty=text_or_none(row[header_index["puesto_trabajo"]]),
                specialty=text_or_none(row[header_index["especialidad"]]),
                assigned_to=text_or_none(row[header_index["personal_asignado"]]),
                raw=raw,
                source_hash=stable_hash(source_payload),
            )
        )

    status_counts = Counter(work_order.status for work_order in work_orders)
    criticality_counts = Counter(work_order.criticality for work_order in work_orders)
    planned_hours = sum(work_order.planned_hours or 0 for work_order in work_orders)
    actual_hours = sum(work_order.actual_hours or 0 for work_order in work_orders)

    return PlansParseResult(
        work_orders=work_orders,
        issues=issues,
        metadata={
            "sheet": sheet.title,
            "sheet_names": sheet_names,
            "header_row": header_row,
            "work_orders": len(work_orders),
            "planned_hours": planned_hours,
            "actual_hours": actual_hours,
            "statuses": dict(status_counts),
            "criticalities": dict(criticality_counts),
            "skipped_rows": skipped_rows,
        },
    )


def to_json(payload: Any) -> str:
    if hasattr(payload, "__dataclass_fields__"):
        payload = asdict(payload)
    return json.dumps(payload, ensure_ascii=False, indent=2, default=str)


def main() -> None:
    parser = argparse.ArgumentParser(description="Parser industrial de Planes Mantencion")
    parser.add_argument("--file", required=True, help="Ruta al archivo Excel de Planes Mantencion")
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument("--dry-run", action="store_true", help="Valida y resume sin escribir datos")
    mode.add_argument("--export", action="store_true", help="Exporta ordenes normalizadas")
    args = parser.parse_args()

    path = Path(args.file).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(path)

    result = parse_plans_file(path)
    if args.export:
        print(to_json(result.export_payload()))
        return
    print(to_json(result.dry_run()))


if __name__ == "__main__":
    main()
